import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from tkinter import *
from tkinter import filedialog, messagebox

def proses_file():
    try:
        # Pilih file input
        file_path = filedialog.askopenfilename(
            title="Pilih File Excel",
            filetypes=[("Excel Files", "*.xlsx *.xls")]
        )
        if not file_path:
            return

        HEADER_ROW_PANDAS = 8   # header = baris ke-9
        USECOLS = "A:P"         # ambil A–P

        # Lokasi simpan
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            title="Simpan File Hasil"
        )
        if not save_path:
            return

        xls = pd.ExcelFile(file_path)
        any_written = False

        with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
            for sheet_name in xls.sheet_names:
                try:
                    df = pd.read_excel(
                        file_path,
                        sheet_name=sheet_name,
                        header=HEADER_ROW_PANDAS,
                        usecols=USECOLS
                    )

                    # Kolom K = kolom ke-11 = index 10
                    if len(df.columns) < 11:
                        print(f"Sheet '{sheet_name}' tidak memiliki kolom K.")
                        continue

                    colK = df.columns[10]

                    # Konversi numeric
                    df[colK] = pd.to_numeric(df[colK], errors="coerce")

                    # Filter K > 10000
                    filtered = df[df[colK] > 10000].copy()
                    if filtered.empty:
                        print(f"Sheet '{sheet_name}' tidak ada data K>10000.")
                        continue

                    # Sort descending
                    filtered.sort_values(by=colK, ascending=False, inplace=True)

                    # Simpan ke sheet baru
                    out_sheet = f"{sheet_name}_filtered"
                    if len(out_sheet) > 31:
                        out_sheet = out_sheet[:31]

                    filtered.to_excel(writer, index=False, sheet_name=out_sheet)
                    any_written = True

                except Exception as e:
                    print(f"Error pada sheet '{sheet_name}': {e}")

        # Tambahkan conditional formatting
        if any_written:
            wb = load_workbook(save_path)

            for ws in wb.worksheets:
                if not ws.title.endswith("_filtered"):
                    continue

                start_row = HEADER_ROW_PANDAS + 2  # header baris 9 → data mulai baris 10
                last_row = ws.max_row

                if last_row < start_row:
                    continue

                # Rumus sesuai permintaan: =K10>10000 (Excel akan ubah otomatis per baris)
                formula = f"$K{start_row}>10000"

                fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                rule = FormulaRule(formula=[formula], stopIfTrue=False, fill=fill)

                # Terapkan highlight ke seluruh baris (A–P)
                ws.conditional_formatting.add(f"A{start_row}:P{last_row}", rule)

            wb.save(save_path)
            messagebox.showinfo("Sukses", f"File selesai diproses:\n{save_path}")
        else:
            messagebox.showinfo("Kosong", "Tidak ada data dengan K > 10000.")

    except Exception as e:
        messagebox.showerror("Error", f"Terjadi kesalahan: {e}")


# GUI Sederhana
if __name__ == "__main__":
    root = Tk()
    root.title("Filter Kolom K > 10000 (Header Baris 9, Kolom A–P)")
    root.geometry("540x200")

    Label(root, text="Filter Kolom K > 10000 (Header = baris 9, Ambil A–P)", 
          font=("Arial", 12, "bold")).pack(pady=14)

    Button(root, text="Pilih File & Proses", font=("Arial", 11), 
           command=proses_file).pack(pady=8)

    Label(root, text="Output: sheet per-sheet bernama Sheet_filtered\n"
                     "Dengan Conditional Formatting rumus =K10>10000",
          wraplength=500, justify="center").pack(pady=6)

    root.mainloop()
