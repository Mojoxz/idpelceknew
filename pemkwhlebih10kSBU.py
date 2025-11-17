import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from tkinter import *
from tkinter import filedialog, messagebox

def proses_file():
    try:
        # Pilih file input
        file_path = filedialog.askopenfilename(
            title="Pilih File Excel",
            filetypes=[("Excel Files", "*.xlsx *.xls")]
        )
        if not file_path:
            return

        # Konfigurasi: header di baris 9 (pandas header=8), gunakan kolom A:P
        HEADER_ROW_PANDAS = 8  # baris ke-9 pada Excel
        USECOLS = "A:P"

        # Pilih lokasi simpan
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            title="Simpan File Hasil"
        )
        if not save_path:
            return

        # Baca semua sheet satu-persatu dan proses
        xls = pd.ExcelFile(file_path)
        any_written = False

        with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
            for sheet_name in xls.sheet_names:
                try:
                    # Baca hanya kolom A:P dengan header di baris ke-9
                    df = pd.read_excel(
                        file_path,
                        sheet_name=sheet_name,
                        header=HEADER_ROW_PANDAS,
                        usecols=USECOLS
                    )

                    # Pastikan kolom N ada (kolom ke-14, index 13, karena A=0)
                    if len(df.columns) < 14:
                        print(f"Sheet '{sheet_name}' kolom kurang dari 14 kolom, dilewati.")
                        continue

                    colN = df.columns[13]  # nama kolom N sesuai header yang dibaca

                    # Konversi kolom N ke numeric
                    df[colN] = pd.to_numeric(df[colN], errors="coerce")

                    # Filter PEMKWH (kolom N) > 10000
                    filtered = df[df[colN] > 10000].copy()

                    if filtered.empty:
                        print(f"Sheet '{sheet_name}': tidak ditemukan nilai {colN} > 10000.")
                        continue

                    # Sort descending berdasarkan kolom N
                    filtered.sort_values(by=colN, ascending=False, inplace=True)

                    # Tambahkan sheet ke output file
                    out_sheet = f"{sheet_name}_filtered"
                    # Pastikan panjang nama sheet <= 31
                    if len(out_sheet) > 31:
                        out_sheet = out_sheet[:31]

                    filtered.to_excel(writer, index=False, sheet_name=out_sheet)
                    any_written = True

                except Exception as e:
                    print(f"Error memproses sheet '{sheet_name}': {e}")

        # Jika ada sheet tertulis, tambah conditional formatting per sheet hasil
        if any_written:
            wb = load_workbook(save_path)

            # Untuk tiap sheet hasil (akhiran "_filtered")
            for ws in wb.worksheets:
                if not ws.title.endswith("_filtered"):
                    continue

                # Tentukan range data
                # Header ada di baris 9 -> data dimulai baris 10 (Excel 1-based)
                start_row = HEADER_ROW_PANDAS + 2  # 8 -> header Excel 9 -> data mulai 10
                last_row = ws.max_row
                if last_row < start_row:
                    continue

                # Terapkan conditional formatting: jika kolom N pada baris >10000 maka highlight A:P
                # Kita gunakan formula yang mengacu ke kolom N pada baris awal range; Excel akan menyesuaikan baris relatif
                formula = f"$N{start_row}>10000"
                fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                rule = FormulaRule(formula=[formula], stopIfTrue=False, fill=fill)

                # Terapkan ke seluruh area A{start_row}:P{last_row} sehingga seluruh baris disorot
                rng = f"A{start_row}:P{last_row}"
                ws.conditional_formatting.add(rng, rule)

            wb.save(save_path)
            messagebox.showinfo("Sukses", f"Hasil tersimpan di:\n{save_path}")
        else:
            messagebox.showinfo("Kosong", "Tidak ada data PEMKWH > 10000 ditemukan pada sheet yang diperiksa.")

    except Exception as e:
        messagebox.showerror("Error", f"Terjadi kesalahan: {e}")


if __name__ == "__main__":
    root = Tk()
    root.title("Filter Kolom N > 10000 (Header Baris 9, Kolom A–P)")
    root.geometry("520x200")

    Label(root, text="Filter Kolom N > 10000 (Header = baris 9, ambil A–P)", font=("Arial", 12, "bold")).pack(pady=14)
    Button(root, text="Pilih File & Proses", font=("Arial", 11), command=proses_file).pack(pady=8)
    Label(root, text="Akan membuat sheet baru untuk tiap sheet hasil: SheetName_filtered", wraplength=480, justify="center").pack(pady=6)

    root.mainloop()
